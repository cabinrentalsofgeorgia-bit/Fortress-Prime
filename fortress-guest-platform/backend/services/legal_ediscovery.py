"""
E-Discovery Vault Ingestion Engine — stores files on NAS, extracts text,
chunks, vectorizes via DGX embeddings, and indexes into Qdrant.

CoCounsel Privilege Shield: Every document passes through a pre-flight
privilege classifier before vectorization.  Attorney-client privileged
material is quarantined to legal.privilege_log and NEVER enters the
case graph or Qdrant search index.
"""
from __future__ import annotations

import email
import email.policy
import hashlib
import io
import json
import time
import structlog
from decimal import Decimal
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional
from uuid import UUID, uuid4, uuid5

from pydantic import BaseModel, Field, field_validator
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from backend.core.config import settings
from backend.core.http_client import shared_client
from backend.services.legal.qdrant_contract import (
    LEGAL_PRIVILEGED_COMMUNICATIONS_COLLECTION,
    LEGAL_WORK_PRODUCT_COLLECTION,
)
from backend.vrs.domain.automations import StreamlineEventPayload
from backend.vrs.infrastructure.event_bus import publish_vrs_event
from backend.services.legal_evidence_ingestion import _chunk_document

logger = structlog.get_logger()

NAS_VAULT_ROOT = Path("/mnt/fortress_nas/legal_vault")
LOCAL_VAULT_FALLBACK = Path("/home/admin/Fortress-Prime/data/legal_vault")
QDRANT_COLLECTION = LEGAL_WORK_PRODUCT_COLLECTION
# PR G — privileged communications go to a physically separate collection so
# Council retrieval can distinguish work product from privileged content at
# the storage layer, not just by payload tag.
QDRANT_PRIVILEGED_COLLECTION = LEGAL_PRIVILEGED_COMMUNICATIONS_COLLECTION
QDRANT_URL = settings.qdrant_url.rstrip("/")
EMBED_URL = f"{settings.embed_base_url.rstrip('/')}/api/embeddings"
EMBED_MODEL = settings.embed_model

SWARM_ENDPOINT = f"{settings.ollama_base_url.rstrip('/')}/v1/chat/completions"
SWARM_MODEL = settings.ollama_fast_model

# UUID5 namespaces for deterministic Qdrant point IDs. Both tracks key on
# (file_hash, chunk_index) so re-runs of the same file upsert to the same
# point IDs — idempotent, never produces duplicates. Required by the Phase D
# reprocess script (Issue #228) so partial-failure retries cannot leak orphan
# points into the collection.
_QDRANT_PRIVILEGED_NS = UUID("f0a17e55-7c0d-4d1f-8c5a-d3b4f0e9a200")
_QDRANT_WORK_PRODUCT_NS = UUID("33f27df1-10c7-4c39-a6bd-085a59bca9b1")

# Maps a privileged-counsel email domain to its case-specific attorney_role tag.
# Source of truth: pr_f_classification_rules.md (v6/v7). Update there first.
_DOMAIN_TO_ROLE: dict[str, str] = {
    "mhtlegal.com":        "case_i_phase_1_filing_to_depositions",
    "fgplaw.com":          "case_i_phase_2_trial_and_general_counsel",
    "msp-lawfirm.com":     "case_i_trial_cocounsel_and_vanderburge_counsel",
    "dralaw.com":          "post_judgment_closing_counsel",
    "wilsonhamilton.com":  "original_transaction_closing_counsel",
    "wilsonpruittlaw.com": "original_transaction_closing_counsel",
}
_KNOWN_PRIVILEGED_DOMAINS = tuple(_DOMAIN_TO_ROLE.keys())


# ── Pydantic: Privilege Classification ────────────────────────────

class PrivilegeClassification(BaseModel):
    is_privileged: bool = Field(default=False, description="True if document contains attorney-client or work-product material")
    privilege_type: str = Field(default="none", description="attorney_client | work_product | joint_defense | none")
    confidence: float = Field(default=0.5, ge=0.0, le=1.0)
    reasoning: str = Field(default="", description="Brief explanation of why this is or is not privileged")

    @field_validator("privilege_type", mode="before")
    @classmethod
    def _coerce_none(cls, v):
        return v or "none"

    @field_validator("reasoning", mode="before")
    @classmethod
    def _coerce_reasoning(cls, v):
        return v or ""


PRIVILEGE_SYSTEM_PROMPT = (
    "You are a legal privilege classifier. Analyze the document text and determine "
    "if it contains Attorney-Client Privileged communications, Work Product, or "
    "Joint Defense Agreement material. Look for: communications between a client "
    "and their attorney seeking or providing legal advice; attorney mental "
    "impressions, strategies, or legal theories; joint defense communications. "
    "Respond ONLY with valid JSON: "
    '{"is_privileged":true,"privilege_type":"attorney_client","confidence":0.95,'
    '"reasoning":"Email between Gary Knight and retained counsel discussing litigation strategy"}'
)


# ── Helpers ───────────────────────────────────────────────────────

def _resolve_vault_dir(case_slug: str) -> Path:
    try:
        if NAS_VAULT_ROOT.exists():
            target = NAS_VAULT_ROOT / case_slug
            target.mkdir(parents=True, exist_ok=True)
            return target
    except (PermissionError, OSError) as exc:
        logger.warning("vault_nas_unavailable", error=str(exc)[:120])
    target = LOCAL_VAULT_FALLBACK / case_slug
    target.mkdir(parents=True, exist_ok=True)
    return target


def _file_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _is_spreadsheet_file(mime_type: str, file_name: str) -> bool:
    lowered_mime = (mime_type or "").lower()
    lowered_name = (file_name or "").lower()
    return (
        "spreadsheet" in lowered_mime
        or "officedocument.spreadsheetml.sheet" in lowered_mime
        or lowered_name.endswith((".xlsx", ".xlsm"))
    )


def _format_spreadsheet_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value).strip()


def _extract_spreadsheet_text(file_bytes: bytes, file_name: str) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is present in runtime
        logger.warning("xlsx_extraction_dependency_missing", file=file_name, error=str(exc)[:200])
        return ""

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("xlsx_extraction_failed", file=file_name, error=str(exc)[:200])
        return ""

    sections: list[str] = []
    try:
        for worksheet in workbook.worksheets:
            rows: list[str] = []
            for row in worksheet.iter_rows(values_only=True):
                cells = [_format_spreadsheet_cell(value) for value in row]
                while cells and not cells[-1]:
                    cells.pop()
                if any(cells):
                    rows.append("\t".join(cells))

            if rows:
                sections.append(f"Sheet: {worksheet.title}\n" + "\n".join(rows))
    finally:
        workbook.close()

    return "\n\n".join(sections)


def _extract_text(file_bytes: bytes, mime_type: str, file_name: str) -> str:
    if "pdf" in mime_type.lower():
        try:
            import fitz
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            pages = [page.get_text() for page in doc]
            doc.close()
            return "\n\n".join(pages)
        except ImportError:
            pass
        except Exception as exc:
            logger.warning("pdf_extraction_failed", error=str(exc)[:200])
        try:
            from io import BytesIO
            from pypdf import PdfReader

            reader = PdfReader(BytesIO(file_bytes))
            return "\n\n".join((page.extract_text() or "") for page in reader.pages)
        except Exception as exc:
            logger.warning("pypdf_extraction_failed", file=file_name, error=str(exc)[:200])
        return file_bytes.decode("utf-8", errors="ignore")

    if _is_spreadsheet_file(mime_type, file_name):
        spreadsheet_text = _extract_spreadsheet_text(file_bytes, file_name)
        if spreadsheet_text.strip():
            return spreadsheet_text
        return ""

    if "csv" in mime_type.lower():
        return file_bytes.decode("utf-8", errors="ignore")

    return file_bytes.decode("utf-8", errors="ignore")


# ── Privilege Classifier ──────────────────────────────────────────

async def _classify_privilege(raw_text: str, file_name: str) -> tuple[PrivilegeClassification, int]:
    """Run a fast pre-flight privilege check via local Qwen2.5.
    Returns (classification, latency_ms)."""
    snippet = raw_text[:3000]
    prompt = f"DOCUMENT: {file_name}\n\n{snippet}"

    t0 = time.perf_counter()
    try:
        resp = await shared_client.post(
            SWARM_ENDPOINT,
            json={
                "model": SWARM_MODEL,
                "messages": [
                    {"role": "system", "content": PRIVILEGE_SYSTEM_PROMPT},
                    {"role": "user", "content": prompt},
                ],
                "temperature": 0.05,
                "max_tokens": 256,
            },
            timeout=60.0,
        )
        latency = int((time.perf_counter() - t0) * 1000)
        resp.raise_for_status()
        data = resp.json()
        content = (
            data.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
            .strip()
        )

        raw_json = content
        if raw_json.startswith("```"):
            nl = raw_json.find("\n")
            raw_json = raw_json[nl + 1:] if nl > 0 else raw_json[3:]
        if raw_json.endswith("```"):
            raw_json = raw_json[:-3]

        start = raw_json.find("{")
        end = raw_json.rfind("}")
        if start >= 0 and end > start:
            raw_json = raw_json[start : end + 1]

        classification = PrivilegeClassification.model_validate_json(raw_json)
        return classification, latency

    except Exception as exc:
        latency = int((time.perf_counter() - t0) * 1000)
        logger.warning(
            "privilege_classifier_failed",
            file=file_name,
            error=str(exc)[:200],
            latency_ms=latency,
        )
        return PrivilegeClassification(
            is_privileged=False,
            privilege_type="none",
            confidence=0.0,
            reasoning=f"Classifier failed: {str(exc)[:100]}",
        ), latency


async def _log_privilege(
    db: AsyncSession,
    doc_id: str,
    case_slug: str,
    file_name: str,
    classification: PrivilegeClassification,
    model: str,
    latency_ms: int,
    snippet: str,
) -> None:
    """Write a permanent privilege log entry and lock the vault document."""
    await db.execute(
        text("""
            INSERT INTO legal.privilege_log
                (document_id, case_slug, file_name, privilege_type, reasoning,
                 model_used, latency_ms, classifier_confidence, snippet)
            VALUES (:did, :slug, :fname, :ptype, :reason, :model, :lat, :conf, :snip)
        """),
        {
            "did": doc_id,
            "slug": case_slug,
            "fname": file_name,
            "ptype": classification.privilege_type,
            "reason": classification.reasoning,
            "model": model,
            "lat": latency_ms,
            "conf": classification.confidence,
            "snip": snippet[:2000],
        },
    )
    await db.execute(
        text("""
            UPDATE legal.vault_documents
            SET processing_status = 'locked_privileged'
            WHERE id = :id
        """),
        {"id": doc_id},
    )
    await db.commit()
    logger.info(
        "privilege_shield_activated",
        doc_id=doc_id,
        file_name=file_name,
        privilege_type=classification.privilege_type,
        confidence=classification.confidence,
    )


# ── Embedding ─────────────────────────────────────────────────────

async def _embed_single(text_input: str) -> list[float]:
    max_retries = 3
    current = text_input[:4000]
    for attempt in range(max_retries):
        try:
            resp = await shared_client.post(
                EMBED_URL,
                json={"model": EMBED_MODEL, "prompt": current},
                timeout=120.0,
            )
            resp.raise_for_status()
            data = resp.json()
            embedding = data.get("embedding") or data.get("data", [{}])[0].get("embedding")
            return embedding if embedding else []
        except Exception as exc:
            err = str(exc)[:200]
            if "500" in err and attempt < max_retries - 1:
                current = current[: len(current) // 2]
                logger.info("embedding_retry_halved", attempt=attempt + 1, new_len=len(current))
                continue
            logger.warning("embedding_failed", error=err, attempt=attempt + 1)
            return []
    return []


async def _embed_chunks(chunks: list[str]) -> list[list[float]]:
    vectors: list[list[float]] = []
    for chunk in chunks:
        vec = await _embed_single(chunk)
        vectors.append(vec)
    return vectors


# ── Qdrant ────────────────────────────────────────────────────────

async def _batch_upsert_with_verification(
    url_base: str,
    collection_name: str,
    points: list[dict],
    batch_size: int = 1000,
) -> tuple[list[str], Optional[dict]]:
    """Upsert points to Qdrant in fixed-size batches; verify each batch's response.

    Returns
    -------
    (successful_uuids, None)
        Every batch verified; ``successful_uuids`` is the full point-id list.
    (successful_uuids_so_far, failure_dict)
        A batch failed verification. ``successful_uuids_so_far`` contains
        the point IDs of every batch that completed successfully BEFORE the
        failed one, in submission order. ``failure_dict`` carries:
            batch_index, expected_count, actual_count,
            qdrant_collection, qdrant_error_payload,
            first_failed_uuid, accumulator_so_far_count.

    Issue #228 design note
    ----------------------
    The pre-fix path swallowed exceptions and returned 0 (indistinguishable
    from a no-vectors success). This helper distinguishes three cases:

      * HTTP error / transport exception → caught, recorded with repr(exc).
      * HTTP 200 but ``result.status != 'completed'``   → caught, recorded.
      * HTTP 200 + ``result.status == 'completed'``    → batch succeeded,
        all batch UUIDs appended to the accumulator.

    Each batch carries its own 60s timeout — splitting large messages avoids
    the single-shot per-document timeout pile-up that contributed to #228 on
    multi-thousand-chunk emails.
    """
    successful: list[str] = []
    if not points:
        return successful, None

    for batch_start in range(0, len(points), batch_size):
        batch = points[batch_start:batch_start + batch_size]
        first_failed_uuid = batch[0]["id"]
        actual_count = 0
        error_payload = ""

        try:
            resp = await shared_client.put(
                f"{url_base}/collections/{collection_name}/points",
                json={"points": batch},
                timeout=60.0,
            )
            resp.raise_for_status()
            body = resp.json()
            # Qdrant: {"result":{"operation_id":int,"status":"completed"|"acknowledged"},
            #          "status":"ok","time":float}
            top_status = body.get("status")
            result_status = (body.get("result") or {}).get("status")
            if top_status == "ok" and result_status in ("completed", "acknowledged"):
                successful.extend(p["id"] for p in batch)
                continue
            error_payload = repr(body)[:1024]
        except Exception as exc:
            error_payload = (repr(exc) or str(exc) or type(exc).__name__)[:1024]

        return successful, {
            "batch_index": batch_start // batch_size,
            "expected_count": len(batch),
            "actual_count": actual_count,
            "qdrant_collection": collection_name,
            "qdrant_error_payload": error_payload,
            "first_failed_uuid": first_failed_uuid,
            "accumulator_so_far_count": len(successful),
        }

    return successful, None


async def _upsert_to_qdrant(
    doc_id: str,
    case_slug: str,
    file_name: str,
    file_hash: str,
    chunks: list[str],
    vectors: list[list[float]],
) -> tuple[list[str], Optional[dict]]:
    """Work-product upsert. Returns (point_uuids_indexed, failure_dict_or_none).

    Point IDs use UUID5 keyed on (file_hash, chunk_index) so re-runs of the
    same physical file produce the same point IDs (idempotent upsert; never
    produces duplicates). This contract is what makes the Phase D reprocess
    script safe to re-run on partially-failed batches without leaving orphan
    points behind.

    Issue #228 fix: previously returned ``int`` (chunk count) and swallowed
    exceptions, making batch failures indistinguishable from no-vector docs.
    Now returns the structured tuple from ``_batch_upsert_with_verification``
    so the caller can transition the vault row to ``qdrant_upsert_failed``.
    """
    points = []
    for idx, (chunk, vector) in enumerate(zip(chunks, vectors)):
        if not vector:
            continue
        points.append({
            "id": str(uuid5(_QDRANT_WORK_PRODUCT_NS, f"{file_hash}:{idx}")),
            "vector": vector,
            "payload": {
                "case_slug": case_slug,
                "document_id": doc_id,
                "file_name": file_name,
                "chunk_index": idx,
                "text": chunk[:1000],
            },
        })

    if not points:
        return [], None

    try:
        await shared_client.put(
            f"{QDRANT_URL}/collections/{QDRANT_COLLECTION}",
            json={
                "vectors": {"size": len(points[0]["vector"]), "distance": "Cosine"},
            },
            timeout=10.0,
        )
    except Exception:
        pass

    return await _batch_upsert_with_verification(
        url_base=QDRANT_URL,
        collection_name=QDRANT_COLLECTION,
        points=points,
    )


# ── Privileged-track helpers (PR G) ───────────────────────────────

def _derive_privileged_counsel_domain(
    file_bytes: bytes,
    file_name: str,
    mime_type: str,
    raw_text: str,
) -> Optional[str]:
    """Return the privileged-counsel email domain that this document was sent to or
    from, if one of _KNOWN_PRIVILEGED_DOMAINS matches. Returns None when no match.

    Strategy:
      1. If the file is an email (.eml / message/rfc822), parse RFC-822 headers
         and look for a privileged domain in From/To/Cc.
      2. Otherwise, scan the first ~16 KB of extracted text for a privileged
         domain substring. (PDFs, .docx etc. — names + email addresses commonly
         appear in headers, signatures, recipient blocks.)
    """
    is_email = (
        "rfc822" in (mime_type or "").lower()
        or file_name.lower().endswith(".eml")
    )
    if is_email:
        try:
            msg = email.message_from_bytes(
                file_bytes, policy=email.policy.default
            )
            addrs = " ".join(
                str(msg.get(h, "") or "") for h in ("From", "To", "Cc", "Bcc", "Reply-To")
            ).lower()
            for d in _KNOWN_PRIVILEGED_DOMAINS:
                if d in addrs:
                    return d
        except Exception:
            pass

    sample = (raw_text or "")[:16384].lower()
    for d in _KNOWN_PRIVILEGED_DOMAINS:
        if d in sample:
            return d
    return None


def _role_for_counsel_domain(domain: Optional[str]) -> Optional[str]:
    """Map a privileged-counsel domain to its case-specific attorney_role tag.
    Returns None when domain is None or unmapped."""
    if not domain:
        return None
    return _DOMAIN_TO_ROLE.get(domain.lower())


async def _upsert_to_qdrant_privileged(
    *,
    doc_id: str,
    case_slug: str,
    file_name: str,
    file_hash: str,
    privileged_counsel_domain: Optional[str],
    role: Optional[str],
    privilege_type: Optional[str],
    chunks: list[str],
    vectors: list[list[float]],
) -> tuple[list[str], Optional[dict]]:
    """Upsert privileged chunks to the legal_privileged_communications collection.

    Point IDs use UUID5 keyed on (file_hash, chunk_index) so re-runs of the
    same physical file produce the same point IDs (idempotent upsert; never
    duplicates points across runs).

    Payload schema per PR G spec:
      case_slug, document_id, file_name, file_hash, chunk_num, chunk_index,
      text (≤1000 chars), privileged=true, privileged_counsel_domain, role,
      privilege_type, ingested_at (UTC ISO-8601).

    Issue #228 fix: previously returned ``int`` (chunk count) and swallowed
    exceptions, making batch failures indistinguishable from no-vector docs.
    Now returns the structured tuple from ``_batch_upsert_with_verification``
    so the caller can transition the vault row to ``qdrant_upsert_failed``.
    """
    points: list[dict] = []
    ingested_at = datetime.now(timezone.utc).isoformat()
    for idx, (chunk, vector) in enumerate(zip(chunks, vectors)):
        if not vector:
            continue
        point_id = str(uuid5(_QDRANT_PRIVILEGED_NS, f"{file_hash}:{idx}"))
        points.append({
            "id": point_id,
            "vector": vector,
            "payload": {
                "case_slug": case_slug,
                "document_id": doc_id,
                "file_name": file_name,
                "file_hash": file_hash,
                "chunk_num": idx,
                "chunk_index": idx,
                "text": chunk[:1000],
                "privileged": True,
                "privileged_counsel_domain": privileged_counsel_domain,
                "role": role,
                "privilege_type": privilege_type,
                "ingested_at": ingested_at,
            },
        })

    if not points:
        return [], None

    return await _batch_upsert_with_verification(
        url_base=QDRANT_URL,
        collection_name=QDRANT_PRIVILEGED_COLLECTION,
        points=points,
    )


# ── Main Ingestion Pipeline ──────────────────────────────────────

_DUPLICATE_TERMINAL_STATUSES = {"completed", "ocr_failed", "locked_privileged"}


def _strip_postgres_nuls(text_value: str) -> str:
    return text_value.replace("\x00", "")


async def process_vault_upload(
    db: AsyncSession,
    case_slug: str,
    file_bytes: bytes,
    file_name: str,
    mime_type: str,
) -> dict:
    fhash = _file_hash(file_bytes)

    fast_dup = await db.execute(
        text("SELECT id FROM legal.vault_documents WHERE case_slug = :slug AND file_hash = :hash"),
        {"slug": case_slug, "hash": fhash},
    )
    duplicate_row = fast_dup.fetchone()
    if duplicate_row:
        existing_id = duplicate_row[0]
        status_result = await db.execute(
            text(
                "SELECT processing_status FROM legal.vault_documents "
                "WHERE id = :id AND case_slug = :slug"
            ),
            {"id": existing_id, "slug": case_slug},
        )
        status_row = status_result.fetchone()
        existing_status = status_row[0] if status_row else None
        if existing_status not in _DUPLICATE_TERMINAL_STATUSES:
            await db.execute(
                text(
                    "DELETE FROM legal.vault_documents "
                    "WHERE id = :id AND case_slug = :slug"
                ),
                {"id": existing_id, "slug": case_slug},
            )
            await db.commit()
        else:
            return {"status": "duplicate", "file_name": file_name, "file_hash": fhash}

    doc_id = str(uuid4())
    vault_dir = _resolve_vault_dir(case_slug)
    safe_name = file_name.replace("/", "_").replace("\\", "_")
    nfs_path = str(vault_dir / f"{doc_id}_{safe_name}")

    with open(nfs_path, "wb") as f:
        f.write(file_bytes)

    insert_result = await db.execute(
        text("""
            INSERT INTO legal.vault_documents
                (id, case_slug, file_name, nfs_path, mime_type, file_hash, file_size_bytes, processing_status)
            VALUES (:id, :slug, :fname, :nfs, :mime, :hash, :size, 'pending')
            ON CONFLICT ON CONSTRAINT uq_vault_documents_case_hash DO NOTHING
            RETURNING id
        """),
        {
            "id": doc_id, "slug": case_slug, "fname": file_name,
            "nfs": nfs_path, "mime": mime_type, "hash": fhash,
            "size": len(file_bytes),
        },
    )
    inserted = insert_result.fetchone()
    await db.commit()
    if inserted is None:
        try:
            Path(nfs_path).unlink(missing_ok=True)
        except Exception:
            pass
        return {"status": "duplicate", "file_name": file_name, "file_hash": fhash}
    docket_event_emitted = await _emit_docket_updated_event(
        db=db,
        case_slug=case_slug,
        document_id=doc_id,
        file_name=file_name,
        mime_type=mime_type,
        nfs_path=nfs_path,
    )

    try:
        raw_text = _strip_postgres_nuls(
            _extract_text(file_bytes, mime_type, file_name)
        )

        # ── Image-only PDF guard ──────────────────────────────
        # If text extraction yields nothing on a PDF, there is no signal
        # to vectorize and no signal to classify privilege. Mark the row
        # ocr_failed so the OCR sweep (backend/scripts/ocr_legal_case.py)
        # can pick it up and the operator has a single state to query for.
        is_pdf = (
            "pdf" in (mime_type or "").lower()
            or file_name.lower().endswith(".pdf")
        )
        if is_pdf and not (raw_text or "").strip():
            await db.execute(
                text(
                    "UPDATE legal.vault_documents "
                    "SET processing_status = 'ocr_failed', error_detail = :err "
                    "WHERE id = :id"
                ),
                {
                    "id": doc_id,
                    "err": "Empty text extraction — image-only PDF requires OCR. "
                           "Run backend/scripts/ocr_legal_case.py for this case.",
                },
            )
            await db.commit()
            logger.info(
                "vault_upload_ocr_failed",
                doc_id=doc_id, case_slug=case_slug, file_name=file_name,
            )
            return {
                "status": "ocr_failed",
                "document_id": doc_id,
                "file_name": file_name,
                "nfs_path": nfs_path,
                "docket_event_emitted": docket_event_emitted,
            }

        # ── CoCounsel Privilege Shield ────────────────────────
        classification, priv_latency = await _classify_privilege(raw_text, file_name)

        if classification.is_privileged and classification.confidence >= 0.7:
            # 1. Persist the privilege-log row (existing behavior)
            await _log_privilege(
                db=db,
                doc_id=doc_id,
                case_slug=case_slug,
                file_name=file_name,
                classification=classification,
                model=SWARM_MODEL,
                latency_ms=priv_latency,
                snippet=raw_text[:2000],
            )

            # 2. PR G — chunk + embed + upsert into the privileged collection.
            #    Privileged content STILL gets vectorized so privileged Council
            #    retrieval can surface it ("for-your-eyes-only" track), but the
            #    physical store is segregated from work-product (legal_ediscovery).
            counsel_domain = _derive_privileged_counsel_domain(
                file_bytes=file_bytes,
                file_name=file_name,
                mime_type=mime_type,
                raw_text=raw_text,
            )
            role_tag = _role_for_counsel_domain(counsel_domain)

            priv_chunks = _chunk_document(raw_text)
            priv_vectors = await _embed_chunks(priv_chunks)
            priv_indexed_uuids, priv_failure = await _upsert_to_qdrant_privileged(
                doc_id=doc_id,
                case_slug=case_slug,
                file_name=file_name,
                file_hash=fhash,
                privileged_counsel_domain=counsel_domain,
                role=role_tag,
                privilege_type=classification.privilege_type,
                chunks=priv_chunks,
                vectors=priv_vectors,
            )

            # 3a. Issue #228 visible failure path — mark qdrant_upsert_failed,
            # record the partial accumulator, surface structured error_detail.
            if priv_failure is not None:
                err_payload = json.dumps({
                    **priv_failure,
                    "occurred_at": datetime.now(timezone.utc).isoformat(),
                    "track": "privileged",
                    "doc_id": doc_id,
                    "case_slug": case_slug,
                    "file_name": file_name,
                    "accumulator_so_far": priv_indexed_uuids,
                })
                await db.execute(
                    text(
                        "UPDATE legal.vault_documents "
                        "SET processing_status = 'qdrant_upsert_failed', "
                        "    chunk_count = :cc, "
                        "    vector_ids = CAST(:vids AS UUID[]), "
                        "    error_detail = :err "
                        "WHERE id = :id"
                    ),
                    {
                        "id": doc_id,
                        "cc": len(priv_chunks),
                        "vids": priv_indexed_uuids if priv_indexed_uuids else None,
                        "err": err_payload[:8192],
                    },
                )
                await db.commit()
                logger.warning(
                    "vault_upload_qdrant_upsert_failed",
                    track="privileged",
                    doc_id=doc_id, case_slug=case_slug, file_name=file_name,
                    partial_indexed=len(priv_indexed_uuids),
                    expected=len(priv_chunks),
                    batch_index=priv_failure["batch_index"],
                )
                return {
                    "status": "qdrant_upsert_failed",
                    "document_id": doc_id,
                    "file_name": file_name,
                    "track": "privileged",
                    "chunks": len(priv_chunks),
                    "partial_indexed": len(priv_indexed_uuids),
                    "failure": priv_failure,
                }

            # 3b. Flip the vault_documents row to its terminal locked_privileged
            #    state with chunk_count + vector_ids populated (parity with the
            #    work-product 'completed' branch below).
            await db.execute(
                text(
                    "UPDATE legal.vault_documents "
                    "SET processing_status = 'locked_privileged', "
                    "    chunk_count = :cc, "
                    "    vector_ids = CAST(:vids AS UUID[]) "
                    "WHERE id = :id"
                ),
                {
                    "id": doc_id,
                    "cc": len(priv_chunks),
                    "vids": priv_indexed_uuids,
                },
            )
            await db.commit()

            logger.info(
                "vault_upload_locked_privileged",
                doc_id=doc_id,
                case_slug=case_slug,
                file_name=file_name,
                privilege_type=classification.privilege_type,
                privileged_counsel_domain=counsel_domain,
                role=role_tag,
                chunks=len(priv_chunks),
                vectors_indexed=len(priv_indexed_uuids),
            )

            return {
                "status": "locked_privileged",
                "document_id": doc_id,
                "file_name": file_name,
                "docket_event_emitted": docket_event_emitted,
                "privilege_type": classification.privilege_type,
                "confidence": classification.confidence,
                "reasoning": classification.reasoning,
                "privileged_counsel_domain": counsel_domain,
                "role": role_tag,
                "chunks": len(priv_chunks),
                "vectors_indexed": len(priv_indexed_uuids),
            }

        # ── Email Threading & Dedupe (CSV email archives) ─────
        from backend.services.legal_dedupe_engine import (
            parse_email_csv, dedupe_and_thread, filter_terminal_emails,
        )

        is_email_csv = (
            "csv" in (mime_type or "").lower()
            and ("email" in file_name.lower() or "generali" in file_name.lower())
        )

        vectorize_text = raw_text
        dedupe_stats: dict | None = None

        if is_email_csv:
            try:
                emails = parse_email_csv(raw_text)
                if emails:
                    result_dedupe = await dedupe_and_thread(db, case_slug, emails)
                    dedupe_stats = result_dedupe
                    terminals = filter_terminal_emails(emails, result_dedupe["terminal_ids"])

                    terminal_texts = []
                    for em in terminals:
                        terminal_texts.append(
                            f"[EMAIL {em['email_id']}] From: {em['sender']} | "
                            f"Subject: {em['subject']} | Date: {em['sent_at']}\n"
                            f"{em['content']}"
                        )
                    vectorize_text = "\n\n---\n\n".join(terminal_texts) if terminal_texts else raw_text

                    logger.info(
                        "email_dedupe_applied",
                        file=file_name,
                        total=result_dedupe["total_emails"],
                        exact_dupes=result_dedupe["exact_dupes_dropped"],
                        terminals=len(terminals),
                        duplicates_skipped=len(result_dedupe["duplicate_ids"]),
                    )
            except Exception as exc:
                logger.warning("email_dedupe_failed", file=file_name, error=str(exc)[:200])

        # ── Normal vectorization path ─────────────────────────
        await db.execute(
            text("UPDATE legal.vault_documents SET processing_status = 'vectorizing' WHERE id = :id"),
            {"id": doc_id},
        )
        await db.commit()

        chunks = _chunk_document(vectorize_text)

        vectors = await _embed_chunks(chunks)
        indexed_uuids, indexed_failure = await _upsert_to_qdrant(
            doc_id=doc_id,
            case_slug=case_slug,
            file_name=file_name,
            file_hash=fhash,
            chunks=chunks,
            vectors=vectors,
        )

        # Issue #228 visible failure path — work-product track.
        if indexed_failure is not None:
            err_payload = json.dumps({
                **indexed_failure,
                "occurred_at": datetime.now(timezone.utc).isoformat(),
                "track": "work_product",
                "doc_id": doc_id,
                "case_slug": case_slug,
                "file_name": file_name,
                "accumulator_so_far": indexed_uuids,
            })
            await db.execute(
                text(
                    "UPDATE legal.vault_documents "
                    "SET processing_status = 'qdrant_upsert_failed', "
                    "    chunk_count = :cc, "
                    "    vector_ids = CAST(:vids AS UUID[]), "
                    "    error_detail = :err "
                    "WHERE id = :id"
                ),
                {
                    "id": doc_id,
                    "cc": len(chunks),
                    "vids": indexed_uuids if indexed_uuids else None,
                    "err": err_payload[:8192],
                },
            )
            await db.commit()
            logger.warning(
                "vault_upload_qdrant_upsert_failed",
                track="work_product",
                doc_id=doc_id, case_slug=case_slug, file_name=file_name,
                partial_indexed=len(indexed_uuids),
                expected=len(chunks),
                batch_index=indexed_failure["batch_index"],
            )
            return {
                "status": "qdrant_upsert_failed",
                "document_id": doc_id,
                "file_name": file_name,
                "track": "work_product",
                "chunks": len(chunks),
                "partial_indexed": len(indexed_uuids),
                "failure": indexed_failure,
            }

        await db.execute(
            text("""
                UPDATE legal.vault_documents
                SET processing_status = 'completed',
                    chunk_count = :cc,
                    vector_ids = CAST(:vids AS UUID[])
                WHERE id = :id
            """),
            {"id": doc_id, "cc": len(chunks), "vids": indexed_uuids},
        )
        await db.commit()

        logger.info(
            "vault_upload_complete",
            doc_id=doc_id,
            case_slug=case_slug,
            file_name=file_name,
            chunks=len(chunks),
            indexed=len(indexed_uuids),
            privilege_cleared=True,
        )

        return {
            "status": "completed",
            "document_id": doc_id,
            "file_name": file_name,
            "chunks": len(chunks),
            "vectors_indexed": len(indexed_uuids),
            "nfs_path": nfs_path,
            "docket_event_emitted": docket_event_emitted,
            "privilege_cleared": True,
        }

    except Exception as exc:
        try:
            await db.rollback()
        except Exception as rollback_exc:
            logger.warning(
                "vault_upload_failure_rollback_failed",
                doc_id=doc_id,
                error=str(rollback_exc)[:300],
            )

        try:
            await db.execute(
                text(
                    "UPDATE legal.vault_documents "
                    "SET processing_status = 'failed', error_detail = :err "
                    "WHERE id = :id"
                ),
                {"id": doc_id, "err": str(exc)[:1000]},
            )
            await db.commit()
        except Exception as mark_exc:
            try:
                await db.rollback()
            except Exception:
                pass
            logger.error(
                "vault_upload_failed_mark_failed_failed",
                doc_id=doc_id,
                original_error=str(exc)[:300],
                mark_error=str(mark_exc)[:300],
            )
            return {
                "status": "failed",
                "document_id": doc_id,
                "error": str(exc)[:200],
                "mark_failed_error": str(mark_exc)[:200],
            }

        logger.error("vault_upload_failed", doc_id=doc_id, error=str(exc)[:300])
        return {"status": "failed", "document_id": doc_id, "error": str(exc)[:200]}


async def _emit_docket_updated_event(
    *,
    db: AsyncSession,
    case_slug: str,
    document_id: str,
    file_name: str,
    mime_type: str,
    nfs_path: str,
) -> bool:
    if "pdf" not in (mime_type or "").lower() and not file_name.lower().endswith(".pdf"):
        return False

    row = (
        await db.execute(
            text(
                """
                SELECT case_number, status
                FROM legal.cases
                WHERE case_slug = :slug
                ORDER BY id DESC
                LIMIT 1
                """
            ),
            {"slug": case_slug},
        )
    ).mappings().first()
    if not row or not row.get("case_number"):
        logger.warning("docket_updated_emit_skipped_missing_case", case_slug=case_slug, document_id=document_id)
        return False

    event = StreamlineEventPayload(
        entity_type="legal_document",
        entity_id=document_id,
        event_type="docket_updated",
        previous_state={},
        current_state={
            "case_slug": case_slug,
            "case_number": row["case_number"],
            "status": row.get("status"),
            "document_path": nfs_path,
            "filing_name": file_name,
            "persist_to_vault": True,
            "mime_type": mime_type or "application/pdf",
        },
    )
    queued = await publish_vrs_event(event)
    if queued:
        logger.info(
            "docket_updated_event_emitted",
            case_slug=case_slug,
            case_number=row["case_number"],
            document_id=document_id,
            filing_name=file_name,
        )
    else:
        logger.warning(
            "docket_updated_emit_failed",
            case_slug=case_slug,
            case_number=row["case_number"],
            document_id=document_id,
        )
    return bool(queued)

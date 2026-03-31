"""
Deterministic FOIA ingestion for Fannin County STR files.
"""
from __future__ import annotations

from collections.abc import Iterable
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
from decimal import Decimal
import io
from pathlib import Path
import re
from typing import Any

from pydantic import BaseModel, ConfigDict, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models.acquisition import (
    AcquisitionIntelEvent,
    AcquisitionOwner,
    AcquisitionProperty,
    FunnelStage,
    SignalSource,
)
from backend.services.acquisition_matching import (
    create_str_signal,
    ensure_property_for_parcel,
    json_safe,
    normalize_address,
    normalize_name,
    resolve_owner,
    resolve_parcel_by_id,
    resolve_property_by_address,
)


class FoiaIngestJobPayload(BaseModel):
    model_config = ConfigDict(extra="forbid")

    spool_path: str
    filename: str
    dry_run: bool = False
    county_name: str = "Fannin"


class FoiaRow(BaseModel):
    model_config = ConfigDict(extra="ignore")

    parcel_id: str | None = None
    owner_legal_name: str | None = None
    tax_mailing_address: str | None = None
    property_address: str | None = None
    primary_residence_state: str | None = None
    fannin_str_cert_id: str | None = None
    blue_ridge_str_permit: str | None = None
    airbnb_listing_id: str | None = None
    vrbo_listing_id: str | None = None
    raw_payload: dict[str, Any] = Field(default_factory=dict)


@dataclass(frozen=True)
class _ColumnAlias:
    canonical: str
    aliases: tuple[str, ...]


_COLUMN_ALIASES: tuple[_ColumnAlias, ...] = (
    _ColumnAlias("parcel_id", ("parcel_id", "parcel id", "parcel", "parcel number", "pin", "apn", "map/parcel")),
    _ColumnAlias("owner_legal_name", ("owner", "owner name", "taxpayer", "name", "owner_legal_name")),
    _ColumnAlias("tax_mailing_address", ("mailing address", "owner mailing address", "tax mailing address", "mail address")),
    _ColumnAlias("property_address", ("property address", "site address", "location address", "rental address")),
    _ColumnAlias("primary_residence_state", ("state", "mailing state", "owner state")),
    _ColumnAlias("fannin_str_cert_id", ("certificate", "certificate number", "excise tax certificate", "str certificate")),
    _ColumnAlias("blue_ridge_str_permit", ("permit", "permit number", "license number", "str permit")),
    _ColumnAlias("airbnb_listing_id", ("airbnb", "airbnb id", "airbnb listing id")),
    _ColumnAlias("vrbo_listing_id", ("vrbo", "vrbo id", "vrbo listing id")),
)


def _normalize_header(value: str) -> str:
    return " ".join(re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).split())


def _normalized_row(row: dict[str, Any]) -> dict[str, Any]:
    return {_normalize_header(str(key)): value for key, value in row.items()}


def _pick_value(row: dict[str, Any], aliases: Iterable[str]) -> str | None:
    for alias in aliases:
        value = row.get(alias)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return None


def _state_from_address(address: str | None) -> str | None:
    text = str(address or "").strip().upper()
    if not text:
        return None
    match = re.search(r"\b([A-Z]{2})\s+\d{5}(?:-\d{4})?\b", text)
    return match.group(1) if match else None


def _decode_csv(file_bytes: bytes) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = file_bytes.decode(encoding)
            reader = csv.DictReader(io.StringIO(text))
            return [dict(row) for row in reader if any(str(value or "").strip() for value in row.values())]
        except UnicodeDecodeError:
            continue
    raise ValueError("Unable to decode FOIA CSV payload.")


def _decode_xlsx(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised in runtime when dependency missing
        raise RuntimeError("openpyxl is required for XLSX FOIA ingestion.") from exc

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    decoded: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        row = {
            headers[index]: raw_row[index]
            for index in range(min(len(headers), len(raw_row)))
            if headers[index]
        }
        if any(str(value or "").strip() for value in row.values()):
            decoded.append(row)
    return decoded


def parse_foia_rows(filename: str, file_bytes: bytes) -> list[FoiaRow]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        rows = _decode_csv(file_bytes)
    elif suffix in {".xlsx", ".xlsm"}:
        rows = _decode_xlsx(file_bytes)
    else:
        raise ValueError("FOIA upload must be CSV or XLSX.")

    parsed_rows: list[FoiaRow] = []
    for raw_row in rows:
        normalized = _normalized_row(raw_row)
        payload: dict[str, Any] = {}
        for alias in _COLUMN_ALIASES:
            payload[alias.canonical] = _pick_value(normalized, alias.aliases)
        payload["raw_payload"] = json_safe(raw_row)
        if not payload["primary_residence_state"]:
            payload["primary_residence_state"] = _state_from_address(payload["tax_mailing_address"])
        parsed_rows.append(FoiaRow.model_validate(payload))
    return parsed_rows


async def run_fannin_foia_ingest(
    db: AsyncSession,
    *,
    filename: str,
    file_bytes: bytes,
    county_name: str = "Fannin",
    dry_run: bool = False,
) -> dict[str, Any]:
    dry_run_savepoint = await db.begin_nested() if dry_run else None
    rows = parse_foia_rows(filename, file_bytes)
    summary = {
        "filename": filename,
        "county_name": county_name,
        "dry_run": dry_run,
        "rows_seen": len(rows),
        "owners_created": 0,
        "properties_promoted": 0,
        "signals_created": 0,
        "intel_events_created": 0,
        "unmatched_rows": 0,
        "warnings": [],
    }

    try:
        for row in rows:
            parcel = await resolve_parcel_by_id(db, row.parcel_id)
            property_match = None
            if parcel is None and row.property_address:
                property_match = await resolve_property_by_address(db, address=row.property_address)
                parcel = property_match.parcel if property_match else None

            owner = await resolve_owner(
                db,
                legal_name=row.owner_legal_name,
                mailing_address=row.tax_mailing_address,
            )
            if owner is None and row.owner_legal_name and row.tax_mailing_address:
                owner = AcquisitionOwner(
                    legal_name=row.owner_legal_name,
                    tax_mailing_address=row.tax_mailing_address,
                    primary_residence_state=row.primary_residence_state,
                )
                db.add(owner)
                await db.flush()
                summary["owners_created"] += 1

            if parcel is None:
                summary["unmatched_rows"] += 1
                summary["warnings"].append(
                    f"Unmatched FOIA row for parcel/address: {row.parcel_id or row.property_address or '<unknown>'}"
                )
                continue

            existing_property = (
                await db.execute(
                    select(AcquisitionProperty).where(AcquisitionProperty.parcel_id == parcel.id).limit(1)
                )
            ).scalar_one_or_none()
            prop = property_match or await ensure_property_for_parcel(db, parcel=parcel, owner=owner)
            if existing_property is None:
                summary["properties_promoted"] += 1
            if prop.owner_id is None and owner is not None:
                prop.owner_id = owner.id
            promoted = False
            if row.fannin_str_cert_id and not prop.fannin_str_cert_id:
                prop.fannin_str_cert_id = row.fannin_str_cert_id
                promoted = True
            if row.blue_ridge_str_permit and not prop.blue_ridge_str_permit:
                prop.blue_ridge_str_permit = row.blue_ridge_str_permit
                promoted = True
            if row.airbnb_listing_id and not prop.airbnb_listing_id:
                prop.airbnb_listing_id = row.airbnb_listing_id
                promoted = True
            if row.vrbo_listing_id and not prop.vrbo_listing_id:
                prop.vrbo_listing_id = row.vrbo_listing_id
                promoted = True
            if prop.pipeline is not None and prop.pipeline.stage == FunnelStage.RADAR:
                prop.pipeline.stage = FunnelStage.TARGET_LOCKED

            await create_str_signal(
                db,
                property_id=prop.id,
                signal_source=SignalSource.FOIA_CSV,
                confidence_score=Decimal("1.00"),
                raw_payload={
                    "filename": filename,
                    "county_name": county_name,
                    "foia_row": row.raw_payload,
                    "parcel_id": row.parcel_id,
                    "property_address": row.property_address,
                },
            )
            summary["signals_created"] += 1

            db.add(
                AcquisitionIntelEvent(
                    property_id=prop.id,
                    event_type="FOIA_STR_SIGNAL",
                    event_description=f"FOIA STR signal ingested from {filename}.",
                    raw_source_data={
                        "filename": filename,
                        "parcel_id": row.parcel_id,
                        "property_address": row.property_address,
                    },
                )
            )
            await db.flush()
            summary["intel_events_created"] += 1

        if dry_run:
            if dry_run_savepoint is not None:
                await dry_run_savepoint.rollback()
        else:
            await db.commit()
        return summary
    except Exception:
        if dry_run_savepoint is not None:
            await dry_run_savepoint.rollback()
        await db.rollback()
        raise
